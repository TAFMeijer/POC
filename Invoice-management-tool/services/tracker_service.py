import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

def get_client_folder(client_name, year):
    """Returns the absolute path to the client's invoice folder for the given year."""
    base_dir = os.path.dirname(os.path.dirname(__file__))
    return os.path.join(base_dir, "invoices", str(year), client_name)

def get_tracker_path(client_name, year):
    folder = get_client_folder(client_name, year)
    return os.path.join(folder, f"Suivi_Factures_{year}.xlsx")

def _initialize_tracker(path):
    """Generates a new Excel file with the correct headers if it doesn't exist."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi Factures"

    headers = ["Date", "N° Facture", "Total (€)", "Payé ?", "Lien Fichier"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="161922", end_color="161922", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Set some column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50

    wb.save(path)

def update_tracker(client_name, year, invoice_data, is_correction=False):
    """
    Updates the client's Excel tracker.
    
    invoice_data expecting:
    {
        "date": "13-04-2026",
        "invoice_number": "2026-001",
        "amount": 120.00,
        "path": "/absolute/path/to/facture.pdf"
    }
    """
    tracker_path = get_tracker_path(client_name, year)
    folder = get_client_folder(client_name, year)
    os.makedirs(folder, exist_ok=True)

    if not os.path.exists(tracker_path):
        _initialize_tracker(tracker_path)

    wb = load_workbook(tracker_path)
    ws = wb.active

    # Check if we should overwrite an existing row based on invoice number
    row_to_overwrite = None
    if is_correction:
        # Scan from bottom to top to find the exact invoice
        for row in range(ws.max_row, 1, -1):
            if ws.cell(row=row, column=2).value == invoice_data["invoice_number"]:
                row_to_overwrite = row
                break

    # Format the data according to columns
    row_data = [
        invoice_data["date"],
        invoice_data["invoice_number"],
        invoice_data["amount"],
        "",  # Paid ? is empty by default so user can fill it
        invoice_data["path"]
    ]

    if is_correction and row_to_overwrite:
        # Overwrite the exact row
        for col_idx, value in enumerate(row_data, start=1):
            # Don't overwrite the 'Paid' status if it was already marked, unless it's a completely new structure?
            # Actually, let's preserve the existing 'Paid?' column (col 4)
            if col_idx == 4:
                continue 
            ws.cell(row=row_to_overwrite, column=col_idx, value=value)
    else:
        # Append new row
        ws.append(row_data)

    wb.save(tracker_path)

def get_latest_invoice_number(client_name, year):
    """Reads the Excel file to find the latest invoice number for this client."""
    tracker_path = get_tracker_path(client_name, year)
    if not os.path.exists(tracker_path):
        return None

    wb = load_workbook(tracker_path)
    ws = wb.active

    # Scan from bottom to top ignoring empty rows
    for row in range(ws.max_row, 1, -1):
        inv_num = ws.cell(row=row, column=2).value
        if inv_num and isinstance(inv_num, str) and inv_num.startswith(f"{year}-"):
            return inv_num

    return None
