"""
Client service — reads client data from an Excel file (clients.xlsx).

Expected columns in the Excel file:
  - name:            Client display name (used for dropdown + calendar matching)
  - title:           Salutation, e.g. "M", "Mme", "Mrs" (optional)
  - contact_person:  Contact person name (for companies)
  - address_line1:   Street address
  - address_line2:   Postal code + city
  - address_line3:   Country (optional)
  - phone:           Phone number (optional)
  - email:           Email address (optional)
  - hourly_rate:     Rate per hour in EUR
  - currency:        Currency code (default EUR)
  - vat_number:      VAT number (empty for individuals)
  - type:            "person" or "company"
"""

import os
import openpyxl

CLIENTS_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "clients.xlsx")


def _load_workbook():
    """Load the clients workbook."""
    if not os.path.exists(CLIENTS_FILE):
        raise FileNotFoundError(
            f"Client database not found at {CLIENTS_FILE}. "
            "Please create a clients.xlsx file — see README for format."
        )
    return openpyxl.load_workbook(CLIENTS_FILE, read_only=True, data_only=True)


def get_all_clients():
    """Return a list of all client names."""
    wb = _load_workbook()
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    name_idx = headers.index("name")

    clients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[name_idx]:
            clients.append(row[name_idx])

    wb.close()
    return sorted(clients)


def get_client(name):
    """Return full client details as a dict for the given name."""
    wb = _load_workbook()
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if row_dict.get("name", "").strip().lower() == name.strip().lower():
            wb.close()
            # Fill in defaults for missing optional fields
            row_dict.setdefault("title", "")
            row_dict.setdefault("contact_person", "")
            row_dict.setdefault("address_line1", "")
            row_dict.setdefault("address_line2", "")
            row_dict.setdefault("address_line3", "")
            row_dict.setdefault("phone", "")
            row_dict.setdefault("email", "")
            row_dict.setdefault("hourly_rate", 0)
            row_dict.setdefault("currency", "EUR")
            row_dict.setdefault("type", "person")
            row_dict.setdefault("billing_type", "calendar")
            row_dict.setdefault("pack_description_template", "Pack de XX heures <br/>A partir du DD/MM/YYYY")
            row_dict.setdefault("pack_hours", 0)
            # Clean None values
            for k, v in row_dict.items():
                if v is None:
                    row_dict[k] = ""
            return row_dict

    wb.close()
    raise ValueError(f"Client '{name}' not found in database.")
