import io
import pandas as pd
import dash
from dash import dcc
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from data_processing import df_w, indicator_order, filter_data


def build_excel_export(n_clicks, region, country, ip, component, exclude_c19rm=False):
    if not country or not ip:
        return dash.no_update

    b_filt, i_filt, w_filt = filter_data(region=region, country=country, ip=ip, component=component, exclude_c19rm=exclude_c19rm)

    # ── SHEET 1: Master Bar Chart Data ─────────────────────────────────────
    b_agg = b_filt.groupby(['Module Parent Component', 'Module'])['Total Amount'].sum().reset_index()

    # Inject Impact/Outcome modules with zero budget so they appear in export
    io_modules = i_filt[i_filt['IndicatorType'].isin(['Impact indicator', 'Outcome indicator'])]['Module'].unique()
    missing = [{'Module Parent Component': str(m).replace(" (Impact/Outcome)", ""),
                'Module': m, 'Total Amount': 0}
               for m in io_modules if m not in b_agg['Module'].values]
    if missing:
        b_agg = pd.concat([b_agg, pd.DataFrame(missing)], ignore_index=True)

    comp_order = {'HIV/AIDS': 1, 'Tuberculosis': 2, 'Malaria': 3, 'RSSH': 4,
                  'Multi-Component': 5, 'Other': 6, 'Program Management': 99}
    b_agg['comp_sort'] = b_agg['Module Parent Component'].map(comp_order).fillna(7)
    b_agg = b_agg.sort_values(by=['comp_sort', 'Total Amount'], ascending=[True, False])

    # Vectorised indicator counts: single groupby instead of 6 len() calls per module
    ind_counts = (i_filt.groupby(['Module', 'IndicatorType', 'IsCustom'])
                  .size().unstack(['IndicatorType', 'IsCustom'], fill_value=0))

    rows_s1 = []
    for _, row in b_agg.iterrows():
        mod = row['Module']
        if pd.isna(mod) or str(mod).strip() == "":
            continue

        def _get(itype, custom):
            try:
                return int(ind_counts.loc[mod, (itype, custom)])
            except (KeyError, TypeError):
                return 0

        rows_s1.append({
            'Component': row['Module Parent Component'],
            'Module': mod,
            'Budget Amount ($M)': round(row['Total Amount'] / 1_000_000, 1),
            'WPTM Count': len(w_filt[w_filt['Module'] == mod]),
            'Coverage (Standard)': _get('Coverage indicator', False),
            'Coverage (Custom)': _get('Coverage indicator', True),
            'Outcome (Standard)': _get('Outcome indicator', False),
            'Outcome (Custom)': _get('Outcome indicator', True),
            'Impact (Standard)': _get('Impact indicator', False),
            'Impact (Custom)': _get('Impact indicator', True),
        })
    df_sheet1 = pd.DataFrame(rows_s1)

    # ── SHEET 2: Budget Tooltip List ───────────────────────────────────────
    group_cols_b = ['Country', 'Module Parent Component', 'Module', 'Intervention']
    if not exclude_c19rm: group_cols_b.append('Source')
    df_sheet2 = (b_filt[group_cols_b + ['Total Amount']]
                 .groupby(group_cols_b)['Total Amount']
                 .sum().reset_index())
    df_sheet2 = df_sheet2[df_sheet2['Total Amount'] != 0]
    df_sheet2['Budget Amount ($M)'] = (df_sheet2['Total Amount'] / 1_000_000).round(1)
    df_sheet2.drop(columns=['Total Amount'], inplace=True)
    df_sheet2 = df_sheet2.sort_values(by=['Module Parent Component', 'Module', 'Budget Amount ($M)'],
                                      ascending=[True, True, False])

    # ── SHEET 3: Indicator Tooltip List ────────────────────────────────────
    cols_i = ['Country', 'Implementation Period Name', 'Module Parent Component',
              'Module', 'IndicatorType', 'IsCustom', 'IndicatorCode',
              'IndicatorCustomName', 'IndicatorDescription']
    if not exclude_c19rm: cols_i.append('Source')

    df_sheet3 = i_filt[cols_i].copy()
    df_sheet3['__sort_code'] = df_sheet3['IndicatorCode'].fillna(df_sheet3['IndicatorCustomName'])
    df_sheet3['__sort_order'] = df_sheet3['__sort_code'].map(indicator_order).fillna(99999)
    df_sheet3['Indicator Name'] = df_sheet3['IndicatorCustomName'].fillna(df_sheet3['IndicatorDescription'])

    group_cols_i = ['Country', 'Module Parent Component', 'Module', 'IndicatorType',
                    'IsCustom', 'IndicatorCode', 'Indicator Name']
    if not exclude_c19rm: group_cols_i.append('Source')
    group_cols_i += ['__sort_order', '__sort_code']

    df_sheet3 = (df_sheet3.groupby(group_cols_i, dropna=False).size().reset_index(name='Count')
                 .sort_values(by=['Module Parent Component', 'Module', 'Count', '__sort_order', '__sort_code'],
                              ascending=[True, True, False, True, True]))
    df_sheet3.drop(columns=['__sort_code', '__sort_order', 'IndicatorCustomName',
                            'IndicatorDescription', 'IndicatorType', 'IsCustom'],
                   inplace=True, errors='ignore')

    # ── SHEET 4: WPTM List ─────────────────────────────────────────────────
    group_cols_w = ['Country', 'Module Parent Component', 'Module', 'KeyActivity']
    if not exclude_c19rm: group_cols_w.append('Source')
    df_sheet4 = (w_filt[group_cols_w]
                 .groupby(group_cols_w, dropna=False)
                 .size().reset_index(name='Count')
                 .sort_values(by=['Module Parent Component', 'Module', 'Count'],
                              ascending=[True, True, False]))

    # ── Write to Excel ─────────────────────────────────────────────────────
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheets = {
            'Master_Chart': df_sheet1,
            'Budget_Tooltips': df_sheet2,
            'Indicator_Tooltips': df_sheet3,
            'WPTM_Tooltips': df_sheet4,
        }
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            for idx, col_name in enumerate(df.columns, 1):
                # 85th-percentile width with 80-char cap
                p85 = df[col_name].astype(str).map(len).quantile(0.85) if not df.empty else 0
                width = min(max(p85, len(str(col_name))) + 1, 80)
                ws.column_dimensions[get_column_letter(idx)].width = width
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

    output.seek(0)
    filename = f"RSSH_Dashboard_Export_{country}_{component}.xlsx"
    return dcc.send_bytes(output.getvalue(), filename)
