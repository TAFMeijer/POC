import pandas as pd
import io
import dash
from dash import dcc
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from data_processing import df_b, df_i, df_w, indicator_order

def build_excel_export(n_clicks, region, country, ip, component):
    if not country or not ip:
        return dash.no_update
        
    # Same filter logic as update_chart
    b_filt = df_b.copy()
    i_filt = df_i.copy()
    w_filt = df_w.copy()
    
    if country != 'ALL':
        b_filt = b_filt[(b_filt['Country'] == country)]
        i_filt = i_filt[(i_filt['Country'] == country)]
        w_filt = w_filt[(w_filt['Country'] == country)]
    elif region and region != 'ALL':
        from data_processing import country_to_region
        b_filt = b_filt[b_filt['Country'].map(country_to_region) == region]
        i_filt = i_filt[i_filt['Country'].map(country_to_region) == region]
        w_filt = w_filt[w_filt['Country'].map(country_to_region) == region]
        
    if ip != 'ALL':
        b_filt = b_filt[(b_filt['Implementation Period Name'] == ip)]
        i_filt = i_filt[(i_filt['Implementation Period Name'] == ip)]
        w_filt = w_filt[(w_filt['Implementation Period Name'] == ip)]
        
    if component != 'ALL':
        b_filt = b_filt[b_filt['Module Parent Component'] == component]
        i_filt = i_filt[i_filt['Module Parent Component'] == component]
        w_filt = w_filt[w_filt['Module Parent Component'] == component]

    # SHEET 1: Master Bar Chart Data
    b_agg = b_filt.groupby(['Module Parent Component', 'Module'])['Total Amount'].sum().reset_index()
    io_modules = i_filt[i_filt['IndicatorType'].isin(['Impact indicator', 'Outcome indicator'])]['Module'].unique()
    for io_mod in io_modules:
        if io_mod not in b_agg['Module'].values:
            parent = str(io_mod).replace(" (Impact/Outcome)", "")
            b_agg = pd.concat([b_agg, pd.DataFrame([{'Module Parent Component': parent, 'Module': io_mod, 'Total Amount': 0}])], ignore_index=True)
            
    comp_order_dict = {
        'HIV/AIDS': 1, 'Tuberculosis': 2, 'Malaria': 3, 'RSSH': 4,
        'Multi-Component': 5, 'Other': 6, 'Program Management': 99
    }
    b_agg['comp_sort'] = b_agg['Module Parent Component'].map(comp_order_dict).fillna(7)
    b_agg = b_agg.sort_values(by=['comp_sort', 'Total Amount'], ascending=[True, False])
    
    rows_s1 = []
    for _, row in b_agg.iterrows():
        mod = row['Module']
        if pd.isna(mod) or mod.strip() == "": continue
        
        im_std = len(i_filt[(i_filt['Module'] == mod) & (i_filt['IndicatorType'] == 'Impact indicator') & (~i_filt['IsCustom'])])
        im_cus = len(i_filt[(i_filt['Module'] == mod) & (i_filt['IndicatorType'] == 'Impact indicator') & (i_filt['IsCustom'])])
        ou_std = len(i_filt[(i_filt['Module'] == mod) & (i_filt['IndicatorType'] == 'Outcome indicator') & (~i_filt['IsCustom'])])
        ou_cus = len(i_filt[(i_filt['Module'] == mod) & (i_filt['IndicatorType'] == 'Outcome indicator') & (i_filt['IsCustom'])])
        co_std = len(i_filt[(i_filt['Module'] == mod) & (i_filt['IndicatorType'] == 'Coverage indicator') & (~i_filt['IsCustom'])])
        co_cus = len(i_filt[(i_filt['Module'] == mod) & (i_filt['IndicatorType'] == 'Coverage indicator') & (i_filt['IsCustom'])])
        
        rows_s1.append({
            'Component': row['Module Parent Component'],
            'Module': mod,
            'Budget Amount ($M)': round(row['Total Amount'] / 1_000_000, 1),
            'WPTM Count': len(w_filt[w_filt['Module'] == mod]),
            'Coverage (Standard)': co_std,
            'Coverage (Custom)': co_cus,
            'Outcome (Standard)': ou_std,
            'Outcome (Custom)': ou_cus,
            'Impact (Standard)': im_std,
            'Impact (Custom)': im_cus,
        })
    df_sheet1 = pd.DataFrame(rows_s1)
    
    # SHEET 2: Budget Tooltip List
    df_sheet2 = b_filt[['Country', 'Module Parent Component', 'Module', 'Intervention', 'Total Amount']].copy()
    # Group identically mapping natively overlapping columns eliminating duplication rows across overlapping Implementation bounds
    df_sheet2 = df_sheet2.groupby(['Country', 'Module Parent Component', 'Module', 'Intervention'])['Total Amount'].sum().reset_index()
    # Explicitly clear pure absolute 0 outputs internally resolving zero-bounds mapping natively
    df_sheet2 = df_sheet2[df_sheet2['Total Amount'] != 0]
    df_sheet2['Budget Amount ($M)'] = (df_sheet2['Total Amount'] / 1_000_000).round(1)
    df_sheet2.drop(columns=['Total Amount'], inplace=True)
    df_sheet2 = df_sheet2.sort_values(by=['Module Parent Component', 'Module', 'Budget Amount ($M)'], ascending=[True, True, False])
    
    # SHEET 3: Indicator Tooltip List
    df_sheet3 = i_filt[['Country', 'Implementation Period Name', 'Module Parent Component', 'Module', 'IndicatorType', 'IsCustom', 'IndicatorCode', 'IndicatorCustomName', 'IndicatorDescription']].copy()
    try:
        df_order = pd.read_excel('data/Indicator order.xlsx')
        indicator_order = dict(zip(df_order['Indicator'], df_order['Order']))
    except Exception:
        indicator_order = {}
    
    df_sheet3['__sort_code'] = df_sheet3['IndicatorCode'].fillna(df_sheet3['IndicatorCustomName'])
    df_sheet3['__sort_order'] = df_sheet3['__sort_code'].map(indicator_order).fillna(99999)
    df_sheet3 = df_sheet3.sort_values(by=['Module Parent Component', 'Module', '__sort_order', '__sort_code', 'Implementation Period Name'])
    
    # Merge custom and standard names natively into a single 'Indicator Name' attribute
    df_sheet3['Indicator Name'] = df_sheet3['IndicatorCustomName'].fillna(df_sheet3['IndicatorDescription'])
    
    # Aggregate to match Tooltips
    group_cols = ['Country', 'Module Parent Component', 'Module', 'IndicatorType', 'IsCustom', 'IndicatorCode', 'Indicator Name', '__sort_order', '__sort_code']
    df_sheet3 = df_sheet3.groupby(group_cols, dropna=False).size().reset_index(name='Count')
    
    df_sheet3 = df_sheet3.sort_values(by=['Module Parent Component', 'Module', 'Count', '__sort_order', '__sort_code'], ascending=[True, True, False, True, True])
    df_sheet3.drop(columns=['__sort_code', '__sort_order', 'IndicatorCustomName', 'IndicatorDescription', 'IndicatorType', 'IsCustom'], inplace=True, errors='ignore')

    # SHEET 4: WPTM List
    df_sheet4 = w_filt[['Country', 'Module Parent Component', 'Module', 'KeyActivity']].copy()
    df_sheet4 = df_sheet4.groupby(['Country', 'Module Parent Component', 'Module', 'KeyActivity'], dropna=False).size().reset_index(name='Count')
    df_sheet4 = df_sheet4.sort_values(by=['Module Parent Component', 'Module', 'Count'], ascending=[True, True, False])

    # Write to Excel BytesIO
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dfs = {
            'Master_Chart': df_sheet1,
            'Budget_Tooltips': df_sheet2,
            'Indicator_Tooltips': df_sheet3,
            'WPTM_Tooltips': df_sheet4
        }
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns, 1):
                if not df.empty:
                    # Use 85th percentile of text length to aggressively size columns, letting outliers wrap 
                    aggressive_len = df[col].astype(str).map(len).quantile(0.85)
                else:
                    aggressive_len = 0
                    
                max_len = max(aggressive_len, len(str(col))) + 1
                
                # Cap the maximum width strictly to 80 to prevent any column from overtaking the screen
                max_len = min(max_len, 80) 
                
                worksheet.column_dimensions[get_column_letter(idx)].width = max_len
                
            # Apply wrap text and align middle
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
        
    output.seek(0)
    filename = f"RSSH_Dashboard_Export_{country}_{component}.xlsx"
    return dcc.send_bytes(output.getvalue(), filename)
